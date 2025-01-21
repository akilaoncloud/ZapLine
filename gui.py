from tkinter import mainloop, StringVar, Text, filedialog, messagebox

import ttkbootstrap as tb   # pip install ttkbootsrap
from ttkbootstrap.tooltip import ToolTip
from ttkbootstrap.constants import *

from PIL import Image, ImageTk, UnidentifiedImageError  # pip install pillow

from os import startfile, chdir
from time import time
from datetime import timedelta
import threading
import logging

from settings import *
from edge import Edge

import openpyxl # pip install openpyxl

class GUI:
    def __init__(self) -> None:
        try: # Detects if you´re using the .exe file or running with an IDE
            chdir('_internal/assets')
        except:
            pass

        # Creates a log file, so it doesn't need to open a terminal.
        logging.basicConfig(filename='session.log', filemode='w', encoding='utf-8', level=logging.INFO) # Standard

    def syncWorkbook(self, read):
        try:
            global wb

            wb = openpyxl.load_workbook(SHEET_PATH, read_only=read)

            if read:
                wb.close()

            if not read:
                wb.save(SHEET_PATH)

            return True

        except PermissionError:
            self.issueHandler(SHEET_ERROR)
            return False

    def openSheet(self):
        selected_tab = self.str_tab.get()

        if not selected_tab == DEFAULT_TAB_OPTION:
            self.str_lin.set('')

        self.nm_tab.set_menu(DEFAULT_TAB_OPTION)

        startfile(SHEET_PATH)

    def updateTabs(self, *args):
        if str(self.nm_tab['state']) == NORMAL and self.syncWorkbook(True):
            self.nm_tab.set_menu(None, *wb.sheetnames)

    def checkTabs(self, *args):
        if self.str_tab.get() == DEFAULT_TAB_OPTION:
            self.n_lin['state'] = DISABLED
        else:
            self.n_lin['state'] = NORMAL
            self.checkEntries()

    def checkEntries(self, *args):
        n_lin = self.str_lin.get()
        max_row = wb[self.str_tab.get()].max_row-1

        if n_lin == '':
            return
        elif not n_lin.isdecimal():
            self.str_lin.set(1)
        elif int(n_lin) < 1:
            self.str_lin.set(1)
        elif int(n_lin) > max_row:
            self.str_lin.set(max_row)

    def checkImage(self):
        if self.path:
            self.rd_img['state'] = NORMAL
            self.rd_msg_img['state'] = NORMAL
        else:
            self.rd_img['state'] = DISABLED
            self.rd_msg_img['state'] = DISABLED

    def insertImage(self):
        try:
            self.path = filedialog.askopenfile().name

            file = Image.open(self.path)

            image_ratio = file.size[0] / file.size[1]
            canvas_ratio = self.canvas.winfo_width() / self.canvas.winfo_height()
            
            if canvas_ratio > image_ratio: # Image is wider than the canvas
                image_h = int(self.canvas.winfo_height())
                image_w = int(image_h * image_ratio)
            else: # Image is taller than canvas
                image_w = int(self.canvas.winfo_width())
                image_h = int(image_w / image_ratio)

            resized_image = file.resize((image_w, image_h))

            self.canvas_image = ImageTk.PhotoImage(resized_image)

            self.lbl_img.destroy() # Deletes the text, and adds an image
            self.canvas.create_image(self.canvas.winfo_width() / 2, self.canvas.winfo_height() / 2, image = self.canvas_image)

            self.checkImage()

        except AttributeError:
            return None
        
        except UnidentifiedImageError:
            self.issueHandler(FILE_IMAGE_ERROR)
            return None    

    def scaler(self, *args):
        self.speed = float('%.1f' % self.scale.get())
        self.n_speed.set(f'{SPEED}: {self.speed}s')

    def openApp(self):
        global window

        # Creates an window
        window = tb.Window(themename=GUI_THEME)
        window.iconbitmap(ICON_PATH)
        window.iconbitmap(default=ICON_PATH)
        window.title(GUI_TITLE)

        # Adjusts the layout
        window.columnconfigure((0,1,2), weight=1)
        window.rowconfigure((0,1,2), weight=1)

        # Status Variable Created
        self.status = StringVar(value = STATUS_DEFAULT)

        # Textbox - Message
        lbl_box = tb.Label(window, text = LABEL_MESSAGE_BOX) # adds a text or subtitle
        lbl_box.grid(row=0,column=0, pady=10)
        self.msg = Text(window, width=50)
        self.msg.grid(row=1,column=0, sticky=NSEW, padx=15)
        # Button - Sheet file
        self.bt_sheet = tb.Button(window, text = BUTTON_OPEN_SHEET, command = self.openSheet)
        self.bt_sheet.grid(row=2,column=0, pady=10)

        # Canvas - Image preview
        self.canvas = tb.Canvas(window, background='#073642', highlightbackground='#0B5162', highlightthickness=1, autostyle=FALSE)
        self.canvas.grid(row=1,column=1, sticky=NSEW)
        # Label - Image preview
        self.lbl_img = tb.Label(window, text = NO_IMAGE, background='#073642')
        self.lbl_img.grid(row=1,column=1)
        # Button - Add an image
        self.insert_img = tb.Button(window, text = INSERT_IMAGE, command = self.insertImage)
        self.insert_img.grid(row=2,column=1)
        self.path = ''
        
        # Frame - Message configurations
        self.config_msg = tb.Frame(window)

        self.syncWorkbook(True)
    
        self.str_tab = StringVar()
        lbl_tab = tb.Label(self.config_msg, text = LABEL_TAB_DROPDOWN)
        lbl_tab.pack(pady=8)
        self.nm_tab = tb.OptionMenu(self.config_msg, self.str_tab, DEFAULT_TAB_OPTION, *wb.sheetnames)
        self.nm_tab.pack(pady=8, fill='x')

        self.str_lin = StringVar()
        lbl_lin = tb.Label(self.config_msg, text = LABEL_LINE_ENTRY)
        lbl_lin.pack(pady=8)
        self.n_lin = tb.Entry(self.config_msg, textvariable=self.str_lin, state=DISABLED) # In-line textbox
        self.n_lin.pack(pady=8, fill='x')

        ToolTip(self.n_lin, text=N_LIN_TOOLTIP, bootstyle=(WARNING, INVERSE))

        # Choose a mode
        self.mode = StringVar(value=0)

        lbl_mode = tb.Label(self.config_msg, text = SEND_CHOOSE_MODE)
        lbl_mode.pack(pady=8)

        radio_modes = tb.Frame(self.config_msg)

        self.rd_msg = tb.Radiobutton(radio_modes, text = SEND_MESSAGE_MODE, value = 0, variable = self.mode) # Msg button (0)
        self.rd_msg.pack(pady=8, anchor=W)

        self.rd_img = tb.Radiobutton(radio_modes, text = SEND_IMAGE_MODE, value = 1, variable = self.mode, state=DISABLED) # Img button (1)
        self.rd_img.pack(pady=8, anchor=W)

        self.rd_msg_img = tb.Radiobutton(radio_modes, text = SEND_IMG_MSG_MODE, value = 2, variable = self.mode, state=DISABLED) # Msg + img button (2)
        self.rd_msg_img.pack(pady=8, anchor=W)

        radio_modes.pack()

        # Status placed
        self.whats = tb.Label(self.config_msg, textvariable = self.status)
        self.whats.pack(pady=8)

        # Start and stop
        self.sync = tb.Button(self.config_msg, text = BUTTON_SYNC, command = self.threadSync, state=NORMAL)
        self.sync.pack(pady=8)

        strt_stop = tb.Frame(self.config_msg)

        self.send = tb.Button(strt_stop, text = BUTTON_SEND, command = self.threadSend, state=DISABLED)
        self.send.pack(pady=8, side=LEFT)

        self.stop = tb.Button(strt_stop, text = BUTTON_STOP, command = self.threadStop, state=DISABLED)
        self.stop.pack(pady=8, side=LEFT)

        strt_stop.pack()

        # Speed Scaler
        scale_frame = tb.Frame(self.config_msg)

        self.n_speed = StringVar()
        self.scale = tb.Scale(scale_frame, from_=5, to=0.5, length=200, value=1, command=self.scaler)
        self.scale.pack(pady=8)
        lbl_speed = tb.Label(scale_frame, textvariable = self.n_speed)
        lbl_speed.pack(pady=[0,8])

        scale_frame.pack()
        ToolTip(scale_frame, text=SCALE_TOOLTIP, bootstyle=(WARNING, INVERSE))
        
        self.scaler()

        self.config_msg.grid(row=0,rowspan=3,column=2, padx=25)

        self.nm_tab.bind('<Button-1>', self.updateTabs)
        self.str_tab.trace_add('write', self.checkTabs)
        self.str_lin.trace_add('write', self.checkEntries)

        mainloop()

    def issueHandler(self, error):
        messagebox.showerror(title=ERROR_TITLE, message=error)

    def threadStop(self):
        self.status.set(STATUS_STOPPING)
        self.stop['state']=DISABLED
        self.running = False
    
    def threadSync(self):

        def sync():
            self.sync['state']=DISABLED
            self.send['state']=DISABLED

            self.status.set(STATUS_SYNCING)
            sync_result = Edge().syncBrowser()

            if sync_result in (SYNC_ERROR, CONNECTION_ERROR):
                self.status.set(STATUS_ERROR)
                self.issueHandler(sync_result)
            else:
                self.status.set(sync_result)
                self.send['state']=NORMAL

            self.sync['state']=NORMAL

        threading.Thread(target=sync).start()

    def threadSend(self):

        def send():
            tab_name = str(self.str_tab.get())

            if not self.syncWorkbook(False):
                return None
            
            try:
                tab = wb[tab_name]
            except KeyError:
                self.issueHandler(NO_TAB_ERROR)
                return None

            self.status.set(STATUS_SENDING)

            self.msg['state']=DISABLED
            self.bt_sheet['state']=DISABLED
            self.insert_img['state']=DISABLED

            self.nm_tab['state']=DISABLED
            self.n_lin['state']=DISABLED
            self.rd_msg['state']=DISABLED
            self.rd_img['state']=DISABLED
            self.rd_msg_img['state']=DISABLED

            self.sync['state']=DISABLED
            self.stop['state']=NORMAL
            self.send['state']=DISABLED
            
            lin_start = 2 if self.str_lin.get()=='' else int(self.str_lin.get())+1
            mode = int(self.mode.get())
            # 'line_one.character_zero', 'end_of_text - 1 character (\n)'
            raw_message = str(self.msg.get('1.0','end-1c'))
            message = raw_message.replace('\xa0', ' ').rstrip('\n') # Remove this characters from the end of the string
            path = self.path
            last_search = ''

            self.running = True
            result = STATUS_DONE

            Edge().resetScreen(self.speed)

            # Set variables used to estimate time
            first_loop = True
            estimated_time = STATUS_ESTIMATIVE_CALC
            time_per_contact = []

            # Searches contacts one by one
            for ctt in tab.iter_rows(min_row=lin_start):
                contact = ctt[1].row-1
                contacts = tab.max_row-1

                if not first_loop:
                    time_per_contact.append(int(time() - start_time))

                    estimated_time_per_contact = int(sum(time_per_contact) / len(time_per_contact))
                    contacts_left = contacts - contact + 1

                    estimative = contacts_left * estimated_time_per_contact

                    estimated_time = timedelta(seconds=estimative)

                self.status.set(f'{STATUS_SENDING} ({contact}/{contacts})\n{STATUS_ESTIMATIVE_LABEL} {estimated_time}')

                start_time = time()

                if self.running:
                    last_search = Edge().sendContact(last_search, ctt, mode, message, path, self.speed)
                    wb.save(SHEET_PATH)

                    if last_search in (DEFAULT_ERROR, CONNECTION_ERROR):
                        result = STATUS_ERROR
                        self.issueHandler(last_search)
                        break
                    
                else:
                    result = STATUS_STOP
                    Edge().resetScreen()
                    break

                first_loop = False

            self.status.set(result)

            # Tries to reset the sheet to keep errors away
            wb.save(SHEET_PATH)
            self.syncWorkbook(True)
            # The errors only occur when the sheet is not closed before sending the messages

            self.msg['state']=NORMAL
            self.bt_sheet['state']=NORMAL
            self.insert_img['state']=NORMAL

            self.nm_tab['state']=NORMAL
            self.n_lin['state']=NORMAL
            self.rd_msg['state']=NORMAL
            self.checkImage()

            self.stop['state']=DISABLED
            self.sync['state']=NORMAL
            self.send['state']=NORMAL

        threading.Thread(target=send).start()  
        
GUI().openApp()